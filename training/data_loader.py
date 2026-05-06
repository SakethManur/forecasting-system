"""Dataset loading and validation utilities."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from utils.config import settings
from utils.exceptions import DataValidationError


def load_sales_excel(file_path: str | Path | None = None) -> pd.DataFrame:
    """Load sales data from an Excel file and normalize column names.

    The loader first honors configured column names, then falls back to
    automatic detection for uploaded datasets with common retail naming.
    """

    path = Path(file_path) if file_path else settings.default_excel_path
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    df = _read_excel(path)
    column_map = detect_sales_columns(df)
    selected_columns = [column_map["date"], column_map["state"], column_map["sales"]]
    if column_map.get("holiday_flag"):
        selected_columns.append(column_map["holiday_flag"])

    df = df[selected_columns].rename(
        columns={source: target for target, source in column_map.items() if source}
    )
    if "holiday_flag" not in df.columns:
        df["holiday_flag"] = 0

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["state"] = df["state"].astype(str).str.strip()
    df["sales"] = pd.to_numeric(df["sales"], errors="coerce")
    df["holiday_flag"] = (
        pd.to_numeric(df["holiday_flag"], errors="coerce").fillna(0).astype(int)
    )

    if df["date"].isna().any():
        raise DataValidationError("Dataset contains invalid date values.")
    if df["state"].eq("").any():
        raise DataValidationError("Dataset contains blank state values.")

    return df[["date", "state", "sales", "holiday_flag"]].sort_values(
        ["state", "date"]
    ).reset_index(drop=True)


def detect_sales_columns(df: pd.DataFrame) -> dict[str, str | None]:
    """Infer date, state, sales, and optional holiday flag columns."""

    normalized = {_normalize_column(column): column for column in df.columns}
    detected = {
        "date": _detect_date_column(df, normalized),
        "state": _detect_named_column(
            normalized,
            [
                "state",
                "states",
                "province",
                "region",
                "market",
                "location",
                "territory",
            ],
        ),
        "sales": _detect_sales_column(df, normalized),
        "holiday_flag": _detect_named_column(
            normalized,
            ["holiday_flag", "holiday", "is_holiday", "festival_flag", "festive_flag"],
            required=False,
        ),
    }
    missing = [name for name in ["date", "state", "sales"] if not detected[name]]
    if missing:
        available = ", ".join(map(str, df.columns))
        raise DataValidationError(
            f"Could not detect required column(s): {', '.join(missing)}. "
            f"Available columns: {available}"
        )
    return detected


def _read_excel(path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path, engine="openpyxl")
    except SyntaxError:
        return _read_excel_with_openpyxl(path)


def _read_excel_with_openpyxl(path: Path) -> pd.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows)
    return pd.DataFrame(rows, columns=headers)


def prepare_daily_sales(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Return one complete regular time series per state.

    Duplicate state/date rows are aggregated. Missing dates for the inferred
    cadence are inserted, and sales gaps are filled with interpolation plus
    forward/backward fallback.
    """

    required = {"date", "state", "sales", "holiday_flag"}
    if not required.issubset(df.columns):
        raise DataValidationError(f"DataFrame must include {sorted(required)}")

    grouped = (
        df.groupby(["state", "date"], as_index=False).agg(
            sales=("sales", lambda values: values.sum(min_count=1)),
            holiday_flag=("holiday_flag", "max"),
        )
        .sort_values(["state", "date"])
    )

    state_frames: dict[str, pd.DataFrame] = {}
    for state, state_df in grouped.groupby("state"):
        frequency, step_days = infer_frequency(state_df["date"])
        state_df = state_df.set_index("date").sort_index()
        full_index = pd.date_range(state_df.index.min(), state_df.index.max(), freq=frequency)
        state_df = state_df.reindex(full_index)
        state_df.index.name = "date"
        state_df["state"] = state
        state_df["sales"] = (
            state_df["sales"]
            .interpolate(method="time")
            .ffill()
            .bfill()
            .fillna(0.0)
        )
        state_df["holiday_flag"] = state_df["holiday_flag"].fillna(0).astype(int)
        prepared = state_df.reset_index()[["date", "state", "sales", "holiday_flag"]]
        prepared.attrs["frequency"] = frequency
        prepared.attrs["step_days"] = step_days
        state_frames[state] = prepared

    return state_frames


def infer_frequency(dates: pd.Series) -> tuple[str, int]:
    ordered = pd.to_datetime(dates).dropna().sort_values().drop_duplicates()
    if len(ordered) < 2:
        return settings.frequency, 1

    deltas = ordered.diff().dropna().dt.days
    median_days = int(max(1, round(float(deltas.median()))))
    first_day = ordered.iloc[0].day_name()[:3].upper()
    if 5 <= median_days <= 9:
        return f"W-{first_day}", 7
    if 27 <= median_days <= 32:
        return "MS", 30
    return "D", 1


def _normalize_column(column: Any) -> str:
    return (
        str(column)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace(".", "_")
    )


def _detect_named_column(
    normalized: dict[str, str], candidates: list[str], required: bool = True
) -> str | None:
    configured = {
        _normalize_column(settings.date_column): settings.date_column,
        _normalize_column(settings.state_column): settings.state_column,
        _normalize_column(settings.sales_column): settings.sales_column,
    }
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
        if candidate in configured and configured[candidate] in normalized.values():
            return configured[candidate]

    for normalized_name, original_name in normalized.items():
        if any(candidate in normalized_name for candidate in candidates):
            return original_name

    if required:
        return None
    return None


def _detect_date_column(df: pd.DataFrame, normalized: dict[str, str]) -> str | None:
    configured = _normalize_column(settings.date_column)
    if configured in normalized:
        return normalized[configured]

    named = _detect_named_column(
        normalized,
        ["date", "order_date", "invoice_date", "week", "week_start", "period"],
        required=False,
    )
    if named:
        return named

    best_column = None
    best_valid = 0
    for column in df.columns:
        parsed = pd.to_datetime(df[column], errors="coerce")
        valid = int(parsed.notna().sum())
        if valid > best_valid and valid / max(len(df), 1) > 0.8:
            best_column = column
            best_valid = valid
    return best_column


def _detect_sales_column(df: pd.DataFrame, normalized: dict[str, str]) -> str | None:
    configured = _normalize_column(settings.sales_column)
    if configured in normalized:
        return normalized[configured]

    named = _detect_named_column(
        normalized,
        ["sales", "sale", "revenue", "amount", "total_sales", "net_sales"],
        required=False,
    )
    if named:
        return named

    numeric_candidates = []
    for column in df.columns:
        numeric = pd.to_numeric(df[column], errors="coerce")
        valid_ratio = numeric.notna().mean()
        if valid_ratio > 0.8:
            numeric_candidates.append((column, float(np.nanmean(np.abs(numeric)))))
    if not numeric_candidates:
        return None
    numeric_candidates.sort(key=lambda item: item[1], reverse=True)
    return numeric_candidates[0][0]
